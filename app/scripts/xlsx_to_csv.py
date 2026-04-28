#!/usr/bin/env python3
"""
Convierte un xlsx a CSV de manera eficiente.

Usa `python-calamine` (parser Rust, ~5-10x más rápido) si está disponible;
si no, cae a `openpyxl` (modo read-only, streaming).

Auto-detecta la fila del header (los xlsx oficiales traen ~17 filas de metadata previas)
escaneando hasta encontrar todas las columnas requeridas (case-insensitive,
post unaccent + Y→I para tolerar variaciones).

Uso:
    xlsx_to_csv.py <input.xlsx> <output.csv> --required COL1,COL2,COL3 [--sheet NAME] [--max-scan-rows 50]

Salida (stdout, una línea):
    OK rows=N header_row=N cols=N elapsed=S.SSs engine=calamine|openpyxl

Códigos de salida:
    0  OK
    1  archivo no existe / engine no disponible / error de parseo
    2  no se encontró el header
"""
import argparse
import csv
import sys
import time
import unicodedata
from pathlib import Path


def normalize_token(s) -> str:
    """unaccent + upper + trim + Y→I (igual que en SQL)."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    nfd = unicodedata.normalize("NFD", s)
    s = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    return s.replace("Y", "I")


def stringify(c):
    """Convierte una celda (str/int/float/datetime/None) a string limpio."""
    if c is None:
        return ""
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    return str(c)


def iter_rows_calamine(path, sheet_name):
    """generator (i, [cells]) usando python-calamine."""
    from python_calamine import CalamineWorkbook

    wb = CalamineWorkbook.from_path(str(path))
    sheet = wb.get_sheet_by_name(sheet_name) if sheet_name else wb.get_sheet_by_index(0)
    if hasattr(sheet, "iter_rows"):
        for i, row in enumerate(sheet.iter_rows(), start=1):
            yield i, list(row)
    else:
        for i, row in enumerate(sheet.to_python(), start=1):
            yield i, list(row)


def iter_rows_openpyxl(path, sheet_name):
    """generator (i, [cells]) usando openpyxl read-only."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            yield i, list(row)
    finally:
        wb.close()


def pick_engine():
    """Devuelve (name, iter_func) prefiriendo calamine."""
    try:
        import python_calamine  # noqa: F401
        return "calamine", iter_rows_calamine
    except ImportError:
        pass
    try:
        import openpyxl  # noqa: F401
        return "openpyxl", iter_rows_openpyxl
    except ImportError:
        return None, None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--required", required=True)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--max-scan-rows", type=int, default=50)
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.is_file():
        print(f"ERROR: archivo no existe: {in_path}", file=sys.stderr)
        return 1

    engine, iter_func = pick_engine()
    if engine is None:
        print("ERROR: ni python-calamine ni openpyxl disponibles", file=sys.stderr)
        return 1

    required_norm = [normalize_token(r) for r in args.required.split(",") if r.strip()]
    started = time.time()

    headers = None
    header_row = None
    rows_written = 0

    try:
        with open(args.output, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)

            for i, raw_row in iter_func(in_path, args.sheet):
                cells = [stringify(c) for c in raw_row]

                if headers is None:
                    if i > args.max_scan_rows:
                        print(
                            f"ERROR: header no encontrado en las primeras {args.max_scan_rows} filas. "
                            f"Requeridas: {args.required}",
                            file=sys.stderr,
                        )
                        return 2
                    upper_set = {normalize_token(c) for c in cells if c.strip()}
                    if all(req in upper_set for req in required_norm):
                        headers = [c.strip() for c in cells]
                        header_row = i
                        while headers and headers[-1] == "":
                            headers.pop()
                        writer.writerow(headers)
                    continue

                if not any(c.strip() for c in cells):
                    continue

                if len(cells) < len(headers):
                    cells = cells + [""] * (len(headers) - len(cells))
                elif len(cells) > len(headers):
                    cells = cells[: len(headers)]

                cells = [c.strip() if isinstance(c, str) else c for c in cells]
                writer.writerow(cells)
                rows_written += 1
    except Exception as e:
        print(f"ERROR engine={engine}: {e}", file=sys.stderr)
        return 1

    if headers is None:
        return 2

    elapsed = time.time() - started
    print(f"OK rows={rows_written} header_row={header_row} cols={len(headers)} elapsed={elapsed:.2f}s engine={engine}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
